import os
import json
import datetime
from flask import Flask, request, render_template, jsonify
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
EXCEL_FILE = 'messages.xlsx'

def save_transactions(data):
    # 從傳入資料中取得交易資料，預期為一個列表
    transactions = data.get("transactions", [])
    # 取得分類及付款方式資料
    categories = data.get("categories", [])
    payment_methods = data.get("paymentMethods", [])
    
    # 建立對照字典： id -> name
    category_map = {cat.get("id", ""): cat.get("name", "") for cat in categories}
    payment_method_map = {pm.get("id", ""): pm.get("name", "") for pm in payment_methods}
    
    # 如果 Excel 檔案存在則讀取，不存在則新建
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
    
    # 處理每一筆交易
    for transaction in transactions:
        date_str = transaction.get("date", "")
        # 使用 categoryId 對應到記帳 web 上的分類名稱
        category_id = transaction.get("categoryId", "")
        category = category_map.get(category_id, "")
        # 使用 paymentMethodId 對應到付款方式名稱
        pm_id = transaction.get("paymentMethodId", "")
        payment_method = payment_method_map.get(pm_id, "")
        amount = transaction.get("amount", "")
        
        # 若沒有日期則跳過
        if not date_str:
            continue
        try:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue
        # 決定工作表名稱，格式為 YYYY-MM
        sheet_name = dt.strftime("%Y-%m")
        
        # 若已有該工作表則取用，否則建立新的工作表並加上標題
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["日期", "分類", "付款方式", "金額"])
        ws.append([date_str, category, payment_method, amount])
    
    wb.save(EXCEL_FILE)

@app.route('/autosave', methods=['POST'])
def autosave():
    data = request.get_json()
    if not data:
        return jsonify({"status": "failed", "message": "沒有接收到資料"}), 400
    save_transactions(data)
    return jsonify({"status": "success", "message": "資料已存入Excel"})

# 原有的 /send_message 路由等
@app.route('/send_message', methods=['POST'])
def send_message():
    if request.is_json:
        data = request.get_json()
        name = data.get('name', '')
        email = data.get('email', '')
        message = data.get('message', '')
    else:
        name = request.form.get('name')
        email = request.form.get('email')
        message = request.form.get('message')
    print(f"收到訊息：{name}, {email}, {message}")
    
    # 儲存至 Excel (範例：追加一筆)
    # ※你可以繼續保留原有 save_to_excel 邏輯
    # save_to_excel(name, email, message)
    
    return jsonify({"status": "success", "message": "訊息已送出！感謝您的聯繫。"})

@app.route('/static/<path:path>')
def static_proxy(path):
    return app.send_static_file(path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/achievements')
def achievements():
    return render_template('achievements.html')

@app.route('/photos')
def photos():
    return render_template('photos.html')

@app.route('/research')
def research():
    return render_template('research.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)